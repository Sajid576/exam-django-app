import re
import os
import uuid
import json
import base64
import pandas
import shutil
import requests
from io import BytesIO
from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader

from exam.models import Level

BASE_URL = 'http://34.126.123.64:8000/' #'http://127.0.0.1:8000/'
# fileName = 'edata.xlsx'
# packageSheetName = "Sheet1"

# data = {}
# wb = load_workbook(fileName)
# data = sheet.values
# columns = next(data)[0:]
# df = pd.DataFrame(data, columns=columns)
# json_str = df.to_json(orient="records")
# print(json_str)
# Put your sheet in the loader
errors = []
difficulty = {"easy", "medium", "hard"}
questionType = {"mcq"}
level = set()
level_instances = Level.objects.all()
for instance in level_instances:
    level.add(instance.name.lower())
#level = {"class 1", "class 2",  "class 3",  "class 4", "class 5", "class 6", "class 7", "class 8", "class 9",
#         "class 10", "class 11", "class 12", "bcs", "undergrad"}

def getBase64Image(image):
    buffered = BytesIO()
    image.save(buffered, format=image.format)
    img_str = base64.b64encode(buffered.getvalue())
    filename = str(uuid.uuid4())+"/"+image.format
    return filename+";base64," + img_str.decode('utf-8')

def validateDuration(duration):
    x = re.fullmatch("[0-9]*\s[0-9]+:[0-9]+:[0-9]+", duration)
    return x

def validateBoolean(boolVal):
    if type(boolVal) is bool:
        return boolVal
    val = boolVal.lower()
    if val == "true" or val == "false":
        return val
    return None

def validateQuestionDifficulty(val):
    val = val.lower()
    if val in difficulty:
        return val
    return None

def validateQuestionType(val):
    val = val.lower()
    if val in questionType:
        return val
    return None

def validateLevel(val):
    val = val.lower()
    if val in level:
        return val
    return None
        
def validate(obj, sheetName):
    i = 2
    for package in obj:
        if package['name'] is None:
            msg = "The name field is empty " + " in row "+ str(i) + " of sheet " + sheetName
            errors.append(msg)

        if package['description'] is None:
            package['description'] = ""

        if validateDuration(package['durationDays']) is None:
            msg = "The duration \"" + package['durationDays']+ '\" in row '+ str(i) + " of sheet " + sheetName+ " is not correctly formatted"
            errors.append(msg)
        if len(package['exams']) != int(package['examCount']):
            msg = "The examcount is " + package['examCount']+ '*' + " in row "+ str(i) + " of sheet " + sheetName+ " but in exams sheet there are "+ str(len(package['exams'])) +" exams."
            errors.append(msg)
        i+=1
        
def validateExam(obj, sheetName):
    i = 2
    for exam in obj:
        if validateDuration(exam['duration']) is None:
            msg = "The duration \"" + exam['duration']+ "\" "+ " in row "+ str(i) + " of sheet " + sheetName+ " is not correctly formatted"
            errors.append(msg)
        if len(exam['questions']) != int(exam['totalQuestion']):
            msg = "The number of total question is " + exam['totalQuestion']+ '*' +" in row "+ str(i) + " of sheet " + sheetName+ " but in exams sheet there are "+ str(len(exam['questions'])) +" questions."
            errors.append(msg)

        if exam['isManual'] is None:
            exam['isManual'] = "TRUE"
        val = validateBoolean(exam['isManual'])
        if val is None:
            msg = "The isManual field \"" + exam['isManual']+ "\"  "+ " in row "+ str(i) + " of sheet " + sheetName+ " is neither true nor false"
            errors.append(msg)

        if exam['paidExam'] is None:
            exam['paidExam'] = "FALSE"
        val = validateBoolean(exam['paidExam'])
        if val is None:
            msg = "The paidExam field \"" + exam['paidExam']+ "\" "+ " in row "+ str(i) + " of sheet " + sheetName + " is neither true nor false"
            errors.append(msg)

        if exam['passingPercent'] is None:
            exam['passingPercent'] = "40"

        if exam['negativePercentage'] is None:
            exam['negativePercentage'] = "0"

        if exam['instruction'] is None:
            exam['instruction'] = ""

        if exam['name'] is None:
            msg = "The name field is empty " + " in row "+ str(i) + " of sheet " + sheetName
            errors.append(msg)

        if exam['totalMark'] is None:
            exam['totalMark'] = ""
        i+=1


def validateQuestion(obj, sheetName):
    i = 2
    for question in obj:
        if question['estimatedTime'] is None:
            question['estimatedTime'] = "0 00:00:00"

        if validateDuration(question['estimatedTime']) is None:
            msg = 'The duration "' + question['estimatedTime']+ '" in row '+ str(i) + " of sheet " + sheetName+ " is not correctly formatted"
            errors.append(msg)

        if question['trueFalse'] is None:
            question['trueFalse'] = "FALSE"

        val = validateBoolean(question['trueFalse'])
        if val is None:
            msg = "The trueFalse field \" + question['trueFalse']+ \"  "+ " in row "+ str(i) + " of sheet " + sheetName+ " is neither true nor false"
            errors.append(msg)

        if question['fillBlank'] is None:
            question['fillBlank'] = "FALSE"

        val = validateBoolean(question['fillBlank'])
        if val is None:
            msg = "The fillBlank field \" + question['fillBlank']+ \" "+ " in row "+ str(i) +" of sheet " + sheetName+ " is neither true nor false"
            errors.append(msg)

        if question['explanation'] is None:
            question['explanation'] = ""

        if question['hint'] is None:
            question['hint'] = ""

        if question['statement'] is None:
            msg = "The statement field is empty " + " in row "+ str(i) + " of sheet " + sheetName
            errors.append(msg)
        
        if question['markPerQues'] is None:
            msg = "The markPerQues field is empty " + " in row "+ str(i)+ " of sheet " + sheetName
            errors.append(msg)

        if question['correctAnswer'] is None:
            msg = "The correctAnswer field is empty " + " in row "+ str(i)+ " of sheet " + sheetName
            errors.append(msg)

        if question['option1'] is None:
            msg = "The option1 field is empty " + " in row "+ str(i) + " of sheet " + sheetName
            errors.append(msg)
        
        if question['option2'] is None:
            msg = "The option2 field is empty " + " in row "+ str(i) + " of sheet " + sheetName
            errors.append(msg)

        if question['option3'] is None:
            msg = "The option3 field is empty " + " in row "+ str(i) + " of sheet " + sheetName
            errors.append(msg)

        if question['option4'] is None:
            msg = "The option4 field is empty " + " in row "+ str(i) + " of sheet " + sheetName
            errors.append(msg)

        if question['difficulty'] is None:
            question['difficulty'] = "medium"

        val = validateQuestionDifficulty(question['difficulty'])
        if val is None:
            msg = 'The difficulty field \"' + question['difficulty']+ '*'+ " in row "+ str(i) + " of sheet " + sheetName + " is not correct"
            errors.append(msg)

        if question['questionType'] is None:
            question['questionType'] = "mcq"

        val = validateQuestionType(question['questionType'])
        if val is None:
            msg = "The questionType field \"" + question['questionType']+ " in row "+ str(i) + " of sheet " + sheetName + " is not correct"
            errors.append(msg)
        
        if question['level'] is None:
            msg = "The level field is empty " + " in row "+ str(i)+ " of sheet " + sheetName
            errors.append(msg)
        else:
            val = validateLevel(question['level'])
            if val is None:
                msg = "The level field \"" + question['level']+ " in row "+ str(i)+ " of sheet " + sheetName + " is not correct"
                errors.append(msg)
        
        if question['subject'] is None:
            msg = "The subject field is empty " + " in row "+ str(i)+ " of sheet " + sheetName
            errors.append(msg)

        if question['chapter'] is  not None:
            question['chapter'] = None

        if question['topic'] is not None:
            question['topic'] = None

        i+=1

def questionsImport(fileName, wb, sheetName):
    sheet = wb[sheetName]
    # get the indexed letter for each column
    ColNames = {}
    Current  = 'A'
    for COL in sheet.iter_cols(1, sheet.max_column):
        ColNames[COL[0].value] = Current
        Current = chr(ord(Current) + 1)

    excel_data_df = pandas.read_excel(fileName, sheet_name=sheetName)
    excel_data_df = excel_data_df[excel_data_df.filter(regex='^(?!Unnamed)').columns]
    json_str = excel_data_df.to_json(orient="records")
    obj = json.loads(json_str)
    
    
    image_loader = SheetImageLoader(sheet)
    i = 2
    for question in obj:
      cellName = ColNames['image']+ str(i)
      if image_loader.image_in(cellName):
        image = image_loader.get(cellName)
        # path = "questions_image/"+str(i-1)+ ".jpg"
        # if image.mode in ("RGBA", "P"):
        #     image = image.convert("RGB")
        imageStr = getBase64Image(image)
        # image.save(path)
        question['image'] = imageStr
      else:
        question['image'] = ""
      i += 1
    
    validateQuestion(obj,sheetName)
    return obj

def examImport(fileName, wb, sheetName):
    sheet = wb[sheetName]
    # get the indexed letter for each column
    ColNames = {}
    Current  = 'A'
    for COL in sheet.iter_cols(1, sheet.max_column):
        ColNames[COL[0].value] = Current
        Current = chr(ord(Current) + 1)

    excel_data_df = pandas.read_excel(fileName, sheet_name=sheetName, dtype = str)
    excel_data_df = excel_data_df[excel_data_df.filter(regex='^(?!Unnamed)').columns]
    json_str = excel_data_df.to_json(orient="records")
    obj = json.loads(json_str)
    
    i = 2
    for exam in obj:
        questionsArray = questionsImport(fileName, wb, sheet[ColNames['questions']+str(i)].hyperlink.location.split('!')[0])
        exam['questions'] = questionsArray
        i += 1
    validateExam(obj, sheetName)
    return obj

def packageImport(fileName, packageSheetName):
    wb = load_workbook(fileName)
    try:
        sheet = wb[packageSheetName]
    except:
        return
    # get the indexed letter for each column
    ColNames = {}
    Current  = 'A'
    for COL in sheet.iter_cols(1, sheet.max_column):
        ColNames[COL[0].value] = Current
        Current = chr(ord(Current) + 1)

    excel_data_df = pandas.read_excel(fileName, sheet_name=packageSheetName,dtype=str)
    excel_data_df = excel_data_df[excel_data_df.filter(regex='^(?!Unnamed)').columns]
    json_str = excel_data_df.to_json(orient="records")
    obj = json.loads(json_str)

    
    image_loader = SheetImageLoader(sheet)

    # data['packages'] = obj
    # row number
    i = 2
    for package in obj:
        imageStr = ""
        cellName = ColNames['image']+ str(i)
        if image_loader.image_in(cellName):
            image = image_loader.get(cellName)
            # path = "packages_image/"+str(i-1)+ ".jpg"
            if image.mode in ("RGBA", "P"):
                image = image.convert("RGB")
            imageStr = getBase64Image(image)
        #   image.save(path)
        
        #print(package['name'])
        package['image'] = imageStr
        examsArray = examImport(fileName, wb, sheet[ColNames['exams']+str(i)].hyperlink.location.split('!')[0])
        package['exams'] = examsArray
        i += 1

    validate(obj, packageSheetName)
    allErrors = {}
    if len(errors) > 0:
        allErrors['errors'] = errors.copy()
        errors.clear()
        return allErrors
    packages = {}
    packages['packages'] = obj
    json_formatted_str = json.dumps(packages)
    return packages

class BulkImport():
    def bulkImport(fileName, packageSheetName):
        endpoint = BASE_URL + 'api/exam/import'
        obj = packageImport(fileName, packageSheetName)
        if obj is None:
            return (None, -1)
        
        if 'errors' in obj.keys():
            return (obj['errors'], 0)
        # with open('data.json', 'w') as f:
        #     json.dump(obj, f)
        # headers = {"Authorization": "Bearer eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJqdGkiOiJhOTNjYmFhNjgwZmUxMWViOWMzZDE0NThkMDBmYmY2ZSIsInVzZXJfaWQiOjEsInVzZXJuYW1lIjoiYWRtaW4iLCJleHAiOjE2MTUzMTg4ODEsInRva2VuX3R5cGUiOiJhY2Nlc3MifQ.pyZ-0OEMiTYFW7LLyWLxSX4x8D6dB22KqPIvxWEZuzk"}
        resp = requests.post(endpoint, json=obj)
        resp_dic = json.loads(resp.content.decode('UTF-8'))
        return (resp_dic, 1)

# if __name__ == "__main__":
#     try:
#         fileName = sys.argv[1]
#         packageSheetName = sys.argv[2]
#     except IndexError:
#         arg = None
#     bulkImport(fileName, packageSheetName)